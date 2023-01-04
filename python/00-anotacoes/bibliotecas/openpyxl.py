import openpyxl as op

wb = op.load_workbook('file.xlsx')  # caminho absoluto do arquivo
wb.sheetnames                       # lista com nomes das abas

ws = wb['Sheet1']                   # nome da aba
wb.create_sheet('NewSheet', 0)      # cria uma nova aba no índice (opcional) 0
sheet = wb.active                   # seleciona a aba ativa

ws.column_dimensions['B'].width = 8 # edita a largura da coluna

ws['B5'].value                      # retorna o valor da célula B5
ws.cell(row=1,column=2).value       # também retorna o valor da célula, na linha 1, coluna 2

value_range = ws['A2':'B5']         # retorna tuplas com as células do range informado
for a, b in value_range:            # desempacotamento de tuplas
    print(a.value, b.value)

ws.iter_rows(min_row=1,max_row=3,min_col=1,max_col=2)   # itera pelas linhas, das linhas 1 a 3, colunas 1 a 2; retorna um objeto com tuplas para cada linha
ws.iter_cols(min_row=1,max_row=3,min_col=1,max_col=2)   # valores min e max são opcionais, retorna um objeto com tuplas para cada coluna

ws.rows     # retorna tuplas com todas as linhas
ws.columns  # retorna tuplas com todas as colunas

ws['B5'].value = 9                  # altera ou acrescenta um valor para a célula
ws.cell(row=5,column=2).value = 18  # outra forma, mesmo resultado

for i in range(2,10):                           # iterando em valores e preenchendo nova coluna
    b_col = ws.cell(row=i,column=2).value
    ws.cell(row=i,column=3).value = b_col*2

wb.save('file.xlsx')    # sempre salvar e fechar o arquivo ao final das edições
wb.close()



# styles

from openpyxl.styles import Font, Color

font_style = Font(name='Calibri', size=14, color='0D8377', italic=True, bold=True, shadow=True, underline='single')
# nome de qualquer fonte disponível na planilha; color hexadecimal sem o #; consultar docs para tipos de underline
a4 = ws['A4']
a4.font = font_style    # aplica o estilo da fonte definido na célula definida

for i in range(2, 10):  # aplicar o estilo para células de uma coluna
    ws.cell(row=i, column=3).font = font_style

from openpyxl.styles import PatternFill

fill_pattern = PatternFill(patternType='solid', fgColor='0D8377')
# consultar documentação para lista de patterns; fgColor hexadecimal sem o #
a4.fill = fill_pattern  # aplica o estilo de preenchimento na célula definida

from openpyxl.styles import Border, Side

top = Side(border_style='medium', color='0D8377')
bottom = Side(border_style='double', color='0D8377')
# consultar documentação para lista de border styles; color hexadecimal sem o #
border = Border(top=top, bottom=bottom)
a4.border = border  # aplica o estilo de preenchimento na célula definida

from openpyxl.styles import numbers

a4 = 'texto'    # células de datas e números são inseridas com formato correto, porém textos são inseridos com o formato de número
a4.number_format = numbers.FORMAT_TEXT  # formata o conteúdo para texto



# fórmulas

a4 = '=SUM(B2:B9)'      # adiciona fórmula de soma
a4 = '=AVERAGE(B2:B9)'  # adiciona fórmula de média

d1 = ws['D1'] = 'Balanço após um ano'
for i in range(2,10):
    balanco = ws.cell(row=i, column=2).value
    taxa = ws.cell(row=i, column=3).value
    balanco_final = (balanco*taxa)+balanco
    ws.cell(row=i, column=4).value = balanco_final



# validação de dados

from openpyxl.worksheet.datavalidation import DataValidation

valid_options = '"Not Started, In Progress, Completed"'
rule = DataValidation(type='list', formula1=valid_options, allow_blank=True)
rule.error = 'Your entry is not valid.'
rule.errorTitle = 'Select Option'

sheet.add_data_validation(rule)
rule.add('C2:100')
# cria uma lista suspensa nas células selecionadas, apenas com as opções definidas, e não aceita valores em brancos ou personalizados